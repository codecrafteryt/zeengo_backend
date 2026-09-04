#!/usr/bin/env python3
"""Export ZEENTRAVEL Kitchen.xlsx → prisma/data/kitchen-seed.json"""
from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

DEFAULT_XLSX = Path.home() / "Downloads" / "ZEENTRAVEL Kitchen 🇷🇺🤴🏻.xlsx"
OUT = Path(__file__).resolve().parents[1] / "prisma" / "data" / "kitchen-seed.json"


def clean(s):
    if s is None:
        return None
    if isinstance(s, float) and s.is_integer():
        s = int(s)
    t = str(s).strip()
    if not t or t.lower() in {"coming soon", "n/a", "#n/a", "≈≈", "false", "true"}:
        return None
    return re.sub(r"\s+", " ", t)


def clean_phone(s):
    if s is None:
        return None
    if isinstance(s, float):
        s = str(int(s)) if s == int(s) else str(s)
    elif isinstance(s, int):
        s = str(s)
    t = clean(s)
    if not t:
        return None
    t = re.split(r"[|/]", t)[0]
    digits = re.sub(r"\D", "", t)
    if len(digits) < 7:
        return None
    if digits.startswith("966") and len(digits) == 13 and digits.endswith("0"):
        digits = digits[:-1]
    if digits.startswith("971") and len(digits) == 13 and digits.endswith("0"):
        digits = digits[:-1]
    if digits.startswith("966") and len(digits) >= 12:
        return "+" + digits[:12]
    if digits.startswith("971") and len(digits) >= 12:
        return "+" + digits[:12]
    if digits.startswith("965") and len(digits) >= 11:
        return "+" + digits[:11]
    if digits.startswith("973") and len(digits) >= 11:
        return "+" + digits[:11]
    if digits.startswith("7") and len(digits) == 11:
        return "+" + digits
    if digits.startswith("8") and len(digits) == 11:
        return "+7" + digits[1:]
    if len(digits) >= 10:
        return "+" + digits
    return None


def parse_rate(s):
    if s is None:
        return None
    if isinstance(s, (int, float)):
        return float(s) if s else None
    m = re.search(r"(\d+(?:\.\d+)?)", str(s).replace(",", "").replace("،", ""))
    return float(m.group(1)) if m else None


def norm_name(s):
    t = clean(s)
    if not t:
        return None
    t = re.sub(r"^[\d\.\-\*]+\s*", "", t).strip("⭐| ").strip()
    return t if len(t) >= 2 else None


def main():
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_XLSX
    if not xlsx.exists():
        raise SystemExit(f"Excel not found: {xlsx}")

    wb = load_workbook(xlsx, data_only=True, read_only=True)
    hotels, activities, guides, drivers, clients = {}, {}, {}, {}, {}

    def add_hotel(name, city, location=None, rate=None, phone=None, email=None, sales_email=None, website=None):
        name = norm_name(name)
        if not name:
            return
        city = clean(city) or "Unknown"
        key = f"{name.lower()}|{city.lower()}"
        notes = []
        if location and clean(location):
            notes.append(f"Address: {clean(location)}")
        r = parse_rate(rate)
        if r:
            notes.append(f"Rate ~{int(r)} RUB")
        if website and clean(website):
            notes.append(f"Website: {clean(website)}")
        if sales_email and clean(sales_email):
            notes.append(f"Sales: {clean(sales_email)}")
        prev = hotels.get(key, {})
        hotels[key] = {
            "name": name,
            "type": "hotel",
            "city": city,
            "phone": clean_phone(phone) or prev.get("phone"),
            "email": clean(email) or clean(sales_email) or prev.get("email"),
            "notes": " | ".join(notes) or prev.get("notes"),
        }

    def add_activity(name, city, location=None, price=None, phone=None):
        name = norm_name(name)
        if not name:
            return
        city = clean(city) or "Unknown"
        key = f"{name.lower()}|{city.lower()}"
        notes = []
        if location and clean(location):
            notes.append(f"Address: {clean(location)}")
        p = parse_rate(price)
        if p:
            notes.append(f"Price ~{int(p)} RUB")
        activities[key] = {
            "name": name,
            "type": "activity",
            "city": city,
            "phone": clean_phone(phone),
            "email": None,
            "notes": " | ".join(notes) or None,
        }

    def add_guide(name, city, phone=None, note=None):
        name = norm_name(name)
        if not name or "مرشد" in name or "مسول" in name:
            return
        city = clean(city) or "Unknown"
        guides[f"{name.lower()}|{city.lower()}"] = {
            "name": name,
            "type": "guide",
            "city": city,
            "phone": clean_phone(phone),
            "email": None,
            "notes": clean(note),
        }

    def add_driver(name, vehicle=None, model=None, year=None, notes=None, city=None, whatsapp=None):
        name = norm_name(name)
        if not name or ("سواقين" in name and "فلاد" not in name):
            return
        yr = int(year) if isinstance(year, (int, float)) and 1990 < year < 2035 else None
        drivers[name.lower()] = {
            "fullName": name,
            "vehicleMake": clean(vehicle),
            "vehicleModel": clean(model) or clean(vehicle),
            "vehicleYear": yr,
            "whatsapp": clean_phone(whatsapp),
            "phone": clean_phone(whatsapp),
            "city": clean(city),
            "notes": clean(notes),
        }

    def add_client(name, phone, pax=None, code=None, dates=None, nights=None, sales=None):
        name = clean(name)
        phone = clean_phone(phone)
        if not name or not phone:
            return
        if name.startswith("~"):
            name = name[1:].strip()
        key = re.sub(r"\D", "", phone)
        party = None
        if pax:
            m = re.search(r"(\d+)\s*adult", str(pax), re.I)
            adults = int(m.group(1)) if m else None
            kids = len(re.findall(r"\b\d+\s*y\b", str(pax), re.I))
            if re.search(r"\bchild\b", str(pax), re.I):
                kids = max(kids, 1)
            if adults:
                party = adults + kids
        n = int(nights) if isinstance(nights, (int, float)) else None
        if n is None and nights:
            m = re.search(r"(\d+)", str(nights))
            n = int(m.group(1)) if m else None
        legacy = str(int(code)) if isinstance(code, (int, float)) else clean(str(code)) if code else None
        clients[key] = {
            "fullName": name,
            "phone": phone,
            "nationality": "SA" if phone.startswith("+966") else ("AE" if phone.startswith("+971") else None),
            "legacyCode": legacy,
            "partySize": party,
            "dateRange": clean(dates),
            "nights": n,
            "salesOwner": clean(sales),
            "rawPax": clean(str(pax)) if pax else None,
        }

    ws = wb["ZEENTRAVEL Hotels "]
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if i < 13:
            continue
        cells = list(row)
        for start, city in [(0, "Moscow"), (7, "Sochi"), (14, "Saint Petersburg"), (21, "Murmansk")]:
            name = cells[start] if start < len(cells) else None
            loc = cells[start + 1] if start + 1 < len(cells) else None
            rate = cells[start + 2] if start + 2 < len(cells) else None
            phone = cells[start + 3] if start + 3 < len(cells) else None
            res_email = cells[start + 4] if start + 4 < len(cells) else None
            sales = cells[start + 5] if start + 5 < len(cells) else None
            website = None
            if isinstance(name, str) and (
                name.startswith("http") or (".ru" in name.lower() and " " not in name and len(name) < 60)
            ):
                website, name, loc = name, loc, None
            if name and str(name).lower() not in {"hotel", "website"}:
                add_hotel(name, city, loc, rate, phone, res_email, sales, website)

    def scrape(sheet, city, name_col=3, loc_col=4, rate_col=5, phone_col=6, email_col=7, sales_col=8):
        if sheet not in wb.sheetnames:
            return
        for row in wb[sheet].iter_rows(values_only=True):
            cells = list(row)
            if len(cells) <= name_col or not cells[name_col]:
                continue
            ns = str(cells[name_col])
            if ns.lower() in {"hotel", "website", "search....."} or "فنادق" in ns or "Stars" in ns:
                continue
            add_hotel(
                ns,
                city,
                cells[loc_col] if loc_col < len(cells) else None,
                cells[rate_col] if rate_col < len(cells) else None,
                cells[phone_col] if phone_col < len(cells) else None,
                cells[email_col] if email_col < len(cells) else None,
                cells[sales_col] if sales_col < len(cells) else None,
            )

    scrape("Moscow 🇷🇺", "Moscow")
    scrape("St peter  🇷🇺", "Saint Petersburg")
    scrape("Murmansk 🇷🇺 ", "Murmansk")
    scrape(" Sochi 🇷🇺", "Sochi")
    scrape("KAZAN  🇷🇺", "Kazan", 0, 1, 2, 3, 4, 5)

    for row in wb["فنادق st peter"].iter_rows(values_only=True):
        cells = list(row)
        if len(cells) < 2 or not cells[1]:
            continue
        add_hotel(cells[1], "Saint Petersburg", cells[4] if len(cells) > 4 else None, None, cells[3] if len(cells) > 3 else None, None, None, cells[2] if len(cells) > 2 else None)

    for i, row in enumerate(wb["ZEENTRAVEL  Activites "].iter_rows(values_only=True), 1):
        if i < 12:
            continue
        cells = list(row)
        add_activity(cells[0] if cells else None, "Moscow", cells[1] if len(cells) > 1 else None, cells[2] if len(cells) > 2 else None, cells[3] if len(cells) > 3 else None)
        if len(cells) > 5:
            add_activity(cells[5], "Sochi", cells[9] if len(cells) > 9 else None, cells[8] if len(cells) > 8 else None, cells[6] if len(cells) > 6 else None)
        if len(cells) > 11:
            add_activity(cells[11], "Saint Petersburg", cells[12] if len(cells) > 12 else None, cells[13] if len(cells) > 13 else None, cells[14] if len(cells) > 14 else None)
        if len(cells) > 16:
            add_activity(cells[16], "Murmansk", None, cells[17] if len(cells) > 17 else None, None)

    for name, city, phone in [
        ("Yousif Jo", "Moscow", "+79151440589"),
        ("Mohamed Nour", "Moscow", "+79161922008"),
        ("Mohsen", "Moscow", "+79930990771"),
        ("karam Tammouz", "Moscow", "+79957834825"),
        ("Mahmoud Fadel", "Saint Petersburg", "+79006424055"),
        ("Shando Hamdy", "Saint Petersburg", None),
        ("Oleg Murmask", "Murmansk", "+79216050806"),
        ("ElZ Murmansk", "Murmansk", "+79216441020"),
        ("Alexander", "Murmansk", "+79197307382"),
        ("Ivan", "Sochi", "+79384370414"),
        ("ALaa", "Sochi", "+79180631888"),
    ]:
        add_guide(name, city, phone)

    for i, row in enumerate(wb["ZEENTRAVEL  Car"].iter_rows(values_only=True), 1):
        cells = list(row)
        if 47 <= i <= 55:
            add_driver(cells[0] if cells else None, cells[2] if len(cells) > 2 else None, None, cells[3] if len(cells) > 3 else None, cells[5] if len(cells) > 5 else None, "Moscow")
        for c in cells:
            if isinstance(c, str) and "wa.me/" in c:
                m = re.search(r"wa\.me/(\d+)", c)
                add_driver("Vladimir Sochi", "V-Class", "V-Class", 2023, "Sochi fleet", "Sochi", m.group(1) if m else None)

    for i, row in enumerate(wb["Date 2025 "].iter_rows(values_only=True), 1):
        if i < 6:
            continue
        cells = list(row)
        add_client(cells[2] if len(cells) > 2 else None, cells[3] if len(cells) > 3 else None, cells[4] if len(cells) > 4 else None, cells[5] if len(cells) > 5 else None, cells[0] if cells else None, cells[6] if len(cells) > 6 else None, cells[9] if len(cells) > 9 else None)

    junk = re.compile(r"^(DAY|باقه|برنامج|website|Hotel|Search|فنادق|شقق|اكوخ|⭐|موسكو|سانت|سوتشي|mailto)", re.I)
    hotels = {k: v for k, v in hotels.items() if v["name"] and not junk.search(v["name"]) and "http" not in v["name"].lower()}
    activities = {k: v for k, v in activities.items() if v["name"] and not junk.search(v["name"]) and "http" not in v["name"].lower()}

    out = {
        "source": str(xlsx.name),
        "hotels": list(hotels.values()),
        "activities": list(activities.values()),
        "guides": list(guides.values()),
        "drivers": list(drivers.values()),
        "clients": list(clients.values()),
    }
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print({k: len(v) if isinstance(v, list) else v for k, v in out.items()})
    print("wrote", OUT)


if __name__ == "__main__":
    main()
